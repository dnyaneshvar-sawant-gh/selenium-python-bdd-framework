import openpyxl

class XLUtils:
    @staticmethod
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getCellData(file, sheetName, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def setCellData(file, sheetName, rownum, colnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(file)
